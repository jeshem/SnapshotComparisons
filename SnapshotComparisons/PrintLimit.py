from openpyxl import Workbook
import openpyxl
from datetime import datetime
from datetime import date
import os

class print_limit(object):
    date = datetime.now()
    today = date.strftime("%Y-%m-%d, %H-%M-%S")

    all_limits=[]
    all_old_limits = []

    #the file name the limits will be stored under
    title = 'NS-OCI Limits '
    latest_file = ''

    keys = [
	    'name',
        'description',
        'limit_name',
        'availability_domain',
        'scope_type',
        'limit',
        'used',
        'available'
        ]

    
    def __init__(self, location, limits, separator):
        self.loc = location
        self.all_limits = limits
        self.separator = separator

        self.save_limits()
        self.find_latest_file(self.loc)
        pass

    ##########################################################################
    # Save limits
    ##########################################################################
    def save_limits(self):
        wb = Workbook()
        current_region = ''

        for limit in self.all_limits:

            #if a service is in a different region, create a new worksheet for that region
            if limit['region_name'] != current_region:
                current_region = limit['region_name']
                current_ws = wb.create_sheet(limit['region_name'])
                #reset where the list of services should begin in each worksheet
                row = 2


            #use keys to create column titles
            for key in range(len(self.keys)):
                current_ws.cell(1, key+1).value = self.keys[key]

            #if a row is already populated, go down one row
            if current_ws.cell(row, 1).value != None:
                row+=1

            #enter limit data into the row
            for col in range(len(self.keys)):
                key = self.keys[col]
                current_ws.cell(row,col+1).value = limit[key]

        #remove empty default sheet
        sheet_to_remove = wb.get_sheet_by_name('Sheet')
        wb.remove_sheet(sheet_to_remove)

        #save the workbook. The name of the workbook will include the date and time it is created in the format of 'yyyy-mm-dd, hh-mm-ss'
        self.title = self.title + self.today + ".xlsx"
        wb.save(self.loc + self.separator + self.title)

    ##########################################################################
    # Find latest file
    ##########################################################################
    def find_latest_file(self, loc):
        dirs = os.listdir(self.loc)
        found_a_file = False

        last_date = self.date


        for file in dirs:
            #if the file is a directory, go into it and check for more files
            if os.path.isdir(self.loc + self.separator + file):
                        list = self.find_latest_file(loc + self.separator + file)

            #get the date the limits file was modified/created
            file_create_date = datetime.fromtimestamp(os.path.getctime(loc + self.separator + file))

            #if last_date is still today, set last date to the file's creation date
            if last_date == self.date:
                last_date = file_create_date
            #find the latest limits
            elif file_create_date < self.date and file_create_date > last_date:
                last_date = file_create_date
                self.latest_file = file

        #if not previous limits file can be found, no comparison can be made
        if self.latest_file == '':
            print("No previous record of limits found")
        else:
            self.retrieve_limits()

    ##########################################################################
    # Retrieve limits
    ##########################################################################
    def retrieve_limits(self):
        print ("Comparing " + self.title + " to: " + self.latest_file)
        last_wb = openpyxl.load_workbook(self.loc + self.separator + self.latest_file)
        current_region = ''

        for sheet in last_wb:
            if sheet.title != 'differences':
                for row in range(sheet.max_row-1):
                    old_limit = {}
                    old_limit['region_name'] = sheet.title
                    #extract old limit data by the row
                    for col in range(len(self.keys)):
                        key = self.keys[col]
                        old_limit[key] = sheet.cell(row+2,col+1).value
                    print ("Appending: " + str(old_limit))
                    self.all_old_limits.append(old_limit)

        last_wb.close()
        self.compare_limits()
    
    ##########################################################################
    # Compare limits
    ##########################################################################
    def compare_limits(self):
        wb = openpyxl.load_workbook(self.loc + self.separator + self.title)

        print(len(self.all_limits))
        print(len(self.all_old_limits))

        for thing in range(len(self.all_old_limits)):
            limit_dif = 0
            
            #if a service's position is the same with both self.all_old_limits and self.all_limits, get the difference
            if self.all_old_limits[thing]['limit_name'] == self.all_limits[thing]['limit_name'] and self.all_old_limits[thing]['region_name'] == self.all_limits[thing]['region_name']:
                limit_dif = self.all_limits[thing]['limit'] - self.all_old_limits[thing]['limit']
            
            #else, go down all_limits until you find the correct service
            else:
                for n in range(len(self.all_limits) - thing):
                    offset = thing+n
                    if self.all_limits[offset]['limit_name'] == self.all_old_limits[thing]['limit_name'] and self.all_limits[offset]['region_name'] == self.all_old_limits[thing]['region_name']:
                        limit_dif = self.all_limits[offset]['limit'] - self.all_old_limits[thing]['limit']

            #if a difference is found
            if limit_dif != 0:
                #if a sheet to record differences does not exist, create one
                if 'differences' not in wb:
                    dif_ws = wb.create_sheet('differences')
                    for col in range(len(self.keys)):
                        key = self.keys[col]
                        dif_ws.cell(1, col+1).value = key
                    #change the column name from 'limit' to 'current - last' 
                    dif_ws.cell(1, 6).value = 'current - last'
                    #set where the sheet should begin listing the limit differences
                    row = 2

                #find the first empty row
                while dif_ws.cell(row, 1).value:
                    row+=1


                for col in range(len(self.keys)):
                    key = self.keys[col]
                    dif_ws.cell(row, col+1).value = self.all_limits[thing][key]

                dif_ws.cell(row, 6).value = limit_dif
                dif_ws.cell(row, 9).value = self.all_limits[thing]['region_name']

        wb.save(self.loc + self.separator + self.title)
        wb.close()