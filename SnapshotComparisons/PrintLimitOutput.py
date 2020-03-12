from openpyxl import Workbook
import openpyxl
from datetime import datetime
import os
import time

class limit_output_and_compare(object):
    date = datetime.date(datetime.now())
    today = str(date)

    FRA_limits = []
    PHX_limits = []
    IAD_limits = []

    title = 'NS-OCI Limits '
    latest_file = ''

    keys = [
	    'name',
        'description',
        'limit_name',
        'availability_domain',
        'scope_type',
        'value',
        'used',
        'available',
        'region_name'
    ]

    
    def __init__(self, location, limits):
        self.loc = location

        for things in limits:
            if things['region_name'] == 'eu-frankfurt-1':
                self.FRA_limits.append(things)
            elif things['region_name'] == 'us-phoenix-1':
                self.PHX_limits.append(things)
            elif things['region_name'] == 'us-ashburn-1':
               self.IAD_limits.append(things)

        pass

    def print_limits(self):
        wb = Workbook()
        FRA_ws = wb.active
        FRA_ws.title = 'FRA'
        PHX_ws = wb.create_sheet('PHX')
        IAD_ws = wb.create_sheet('IAD')
        
        for row in range(len(self.FRA_limits)):
            thing = self.FRA_limits[row]
            for col in range(len(self.keys)):
                key = self.keys[col]
                FRA_ws.cell(row+1, col+1).value = thing[key]

        for row in range(len(self.PHX_limits)):
            thing = self.PHX_limits[row]
            for col in range(len(self.keys)):
                key = self.keys[col]
                PHX_ws.cell(row+1, col+1).value = thing[key]

        for row in range(len(self.IAD_limits)):
            thing = self.IAD_limits[row]
            for col in range(len(self.keys)):
                self.key = self.keys[col]
                IAD_ws.cell(row+1, col+1).value = thing[key]

        wb.save(self.loc + "\\" + self.title + self.today + ".xlsx")

    #find new files and insert them into a list
    def find_new_files(self, loc):
        dirs = os.listdir(self.loc)
        found_a_file = False

        last_date = time.time()

        for file in dirs:
            #if the file is a directory, go into it and check for more files
            if os.path.isdir(self.loc + "\\" + file):
                        list = self.find_new_files(loc + "\\" + file)

            #get the date the limits file was modified/created
            file_create_date = os.path.getctime(loc + "\\" + file)
            
            #find the latest limits
            if file_create_date < last_date:
                last_date = file_create_date
                self.latest_file = file

        if self.latest_file == '':
            print("No previous limits found")
        else:
            print(self.latest_file)
            self.compare_limits()

    def compare_limits(self):
        last_wb = openpyxl.load_workbook(self.loc + "\\" + self.latest_file)
        last_FRA_ws = last_wb['FRA']
        last_PHX_ws = last_wb['PHX']
        last_IAD_ws = last_wb['IAD']

        last_FRA_limits = []
        last_PHX_limits = []
        last_IAD_limits = []

        FRA_dif = []
        PHX_dif = []
        IAD_dif = []

        #get the limits from each region in the last limits file
        for row in range(last_FRA_ws.max_row):
            cell_value = last_FRA_ws.cell(row+1, 6).value
            last_FRA_limits.append(cell_value)
        for row in range(last_PHX_ws.max_row):
            cell_value = last_PHX_ws.cell(row+1, 6).value
            last_PHX_limits.append(cell_value)
        for row in range(last_IAD_ws.max_row):
            cell_value = last_IAD_ws.cell(row+1, 6).value
            last_IAD_limits.append(cell_value)

        if len(last_FRA_limits)>1:
            for thing in range(len(last_FRA_limits)):
                difference = self.FRA_limits[thing]['value'] - last_FRA_limits[thing]
                val = {'region_name': self.FRA_limits[thing]['region_name'],
                       'name': self.FRA_limits[thing]['limit_name'],
                       'difference': difference}
                FRA_dif.append(val)
        if len(last_PHX_limits)>1:
            for thing in range(len(last_PHX_limits)):
                difference = self.PHX_limits[thing]['value'] - last_PHX_limits[thing]
                val = {'region_name': self.PHX_limits[thing]['region_name'],
                       'name': self.PHX_limits[thing]['limit_name'],
                       'difference': difference}
                PHX_dif.append(val)
        if len(last_IAD_limits)>1:
            for thing in range(len(last_PHX_limits)):
                difference = self.IAD_limits[thing]['value'] - last_IAD_limits[thing]
                val = {'region_name': self.IAD_limits[thing]['region_name'],
                       'name': self.IAD_limits[thing]['limit_name'],
                       'difference': difference}

        
        for thing in FRA_dif:
            if thing['difference'] != 0:
                print(str(thing))       
        for thing in PHX_dif:
            if thing['difference'] != 0:
                print(str(thing))
        for thing in IAD_dif:
            if thing['difference'] != 0:
                print(str(thing))


        pass

