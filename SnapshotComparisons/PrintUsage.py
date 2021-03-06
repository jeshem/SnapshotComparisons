from openpyxl import Workbook
from datetime import datetime
from datetime import date

class print_usage(object):
    date = datetime.now()
    today = date.strftime("%Y-%m-%d, %H-%M-%S")

    all_usages=[]
    
    #the file name the usages will be stored under
    title = 'NSOCI Usages '

    keys = [
	    'name',
        'description',
        'limit_name',
        'availability_domain',
        'scope_type',
        'limit',
        'used',
        'available'
        ]

    
    def __init__(self, location, usages, separator):
        self.loc = location
        self.all_usages = usages
        self.separator = separator

        self.save_usages()
        pass

    ##########################################################################
    # Save usages
    ##########################################################################
    def save_usages(self):
        wb = Workbook()
        current_region = ''

        for each_usage in self.all_usages:
            for usage in each_usage:
                print(usage)

                #if a service is in a different region, create a new worksheet for that region
                if usage['region_name'] != current_region:
                    current_region = usage['region_name']
                    current_ws = wb.create_sheet(usage['region_name'])
                    #reset where the list of services should begin in each worksheet
                    row = 4

                current_ws.cell(1, 1).value = "compartment_id:"
                current_ws.cell(1, 2).value = usage['compartment_id']
                current_ws.column_dimensions['A'].width = 16.43

                #use keys to create column titles
                for key in range(len(self.keys)):
                    current_ws.cell(3, key+1).value = self.keys[key]

                #if a row is already populated, go down one row
                if current_ws.cell(row, 1).value != None:
                    row+=1

                #enter limit data into the row
                for col in range(len(self.keys)):
                    key = self.keys[col]
                    current_ws.cell(row,col+1).value = usage[key]

        #remove empty default sheet
        sheet_to_remove = wb.get_sheet_by_name('Sheet')
        wb.remove_sheet(sheet_to_remove)

        #save the workbook. The name of the workbook will include the date and time it is created in the format of 'yyyy-mm-dd, hh-mm-ss'
        self.title = self.title + self.today + ".xlsx"
        wb.save(self.loc + self.separator + self.title)

    pass




